#!/usr/bin/env python3

import sys
from openpyxl import load_workbook

def extract_sheet_names(xlsx_path, output_txt_path):
    # Load the workbook in read-only mode
    wb = load_workbook(filename=xlsx_path, read_only=True)
    
    # Get all sheet names
    sheet_names = wb.sheetnames
    
    # Write each sheet name to a new line in the output file
    with open(output_txt_path, "w") as f:
        for name in sheet_names:
            f.write(name + "\n")
    
    print(f"Sheet names have been written to {output_txt_path}")

if __name__ == "__main__":
    # Ensure proper command-line usage
    if len(sys.argv) != 3:
        print("Usage: python script.py <input.xlsx> <output.txt>")
        sys.exit(1)
    
    input_xlsx = sys.argv[1]
    output_txt = sys.argv[2]
    extract_sheet_names(input_xlsx, output_txt)
